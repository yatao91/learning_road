# -*- coding: utf-8 -*-
"""
1.获取待写入excel数据
2.写入数据至excel中
3.上传excel至oss中
"""
from tempfile import NamedTemporaryFile

from elasticsearch import Elasticsearch
from openpyxl import Workbook
import oss2
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.application import MIMEApplication


def get_data():
    """
    获取待写入excel数据
    """
    es = Elasticsearch(hosts="39.104.48.117:9288", timeout=60)
    body = {
        "sort": {"cursor": "desc"},
        "_source": ["publish_time", "title"],
        "size": 5
    }
    resp = es.search(index="data-warehouse-zhongzhao-common-bid-result-v7", doc_type="doc", body=body)
    hits = resp["hits"]["hits"]
    return hits


def write_to_excel(hits: list):
    """
    写入数据至excel中
    """
    wb = Workbook()
    ws = wb.active
    ws.title = 'test'
    ws['A1'] = 'publish_time'
    ws['B1'] = 'title'

    num = 2
    for hit in hits:
        data = hit['_source']
        publish_time = data.get("publish_time")
        title = data.get("title")
        ws[f'A{num}'] = publish_time
        ws[f'B{num}'] = title
        num += 1

    with NamedTemporaryFile() as tmp:
        wb.save(tmp.name)
        tmp.seek(0)
        stream = tmp.read()
        return stream


def upload_to_oss(filename: str, content: bytes):
    """
    上传文件到OSS
    """
    auth = oss2.Auth(
        access_key_id="",
        access_key_secret=""
    )
    bucket = oss2.Bucket(
        auth=auth,
        endpoint="oss-cn-huhehaote.aliyuncs.com",
        bucket_name="zhongzhao-file-download"
    )
    result = bucket.put_object(key=filename, data=content)

    print(bucket.sign_url("GET", filename, 120))


def send_email_with_file(file_stream: bytes, filename: str):
    """
    通过email发送文件
    """
    msg = MIMEMultipart()

    # 构建正文
    part_text = MIMEText("文件下载", _charset="utf-8")
    msg.attach(part_text)  # 把正文添加到邮件中去

    # 构建邮件附件
    part_attch1 = MIMEApplication(file_stream)  # 打开附件
    part_attch1.add_header('Content-Disposition', 'attachment', filename=filename)  # 为附件命名
    msg.attach(part_attch1)  # 添加附件

    msg['Subject'] = "数据下载"
    msg['From'] = '460811852@qq.com'
    msg['To'] = '460811852@qq.com'
    smtp = smtplib.SMTP_SSL("smtp.qq.com", port=465)
    smtp.login("460811852@qq.com", "lcxiehbbtlscbjdb")
    smtp.sendmail("460811852@qq.com", "460811852@qq.com", msg.as_string())
    smtp.quit()


def main():
    hits = get_data()
    stream = write_to_excel(hits)
    send_email_with_file(file_stream=stream, filename="test_excel.xlsx")


if __name__ == '__main__':
    main()
