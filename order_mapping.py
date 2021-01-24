# -*- coding: utf-8 -*-
from openpyxl import load_workbook

wb = load_workbook(filename="test_order.xlsx")
ws = wb.active

order_id_mapping = {}
product_array = []

for row in ws.iter_rows(min_row=2):
    order_id, _, product, _ = row
    order_index = order_id_mapping.get(order_id.value)

    if order_index:
        order_product_array = product_array[order_index]
        order_product_array.append(product.value)
    else:
        order_product_array = [product.value]
        product_array.append(order_product_array)

print(product_array)
import time
